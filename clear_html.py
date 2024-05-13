import re
from openpyxl import load_workbook


def clear_xlsx_from_html(path_to_xlsx_file):
    global cell
    # Load the Excel file
    workbook = load_workbook(path_to_xlsx_file)
    sheet = workbook.active

    # Define a function to remove HTML tags
    def remove_html_tags(text):
        clean = re.compile('<.*?>')
        return re.sub(clean, '', text)

    # Iterate through each cell and remove HTML tags
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=8):
        for cell in row:
            if cell.value:
                cell.value = remove_html_tags(str(cell.value))
    # Save the modified workbook
    workbook.save('your_file_cleaned.xlsx')


if __name__ == '__main__':
    clear_xlsx_from_html('/Users/evgeniy/Desktop/English_save_my_collection.xlsx')
